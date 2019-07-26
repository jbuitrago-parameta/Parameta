import pandas as pd
import os
import numpy as np
import datetime
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import cred_gmail
import time
import re
import xlwt 
from xlwt import Workbook
import openpyxl

os.chdir(r'G:\Mi unidad\Data\Grupo Bios\Automatizacion_Items\Consecutivos SIESA')

dict_comp = {'CONTEGRAL': '001', 'FINCA': '100', 'AVICOLA TRIPLE A': '230', 'PIC': '120', 'OPAV': '020'}
dict_mes = {'1':'ENE','2':'FEB','3':'MAR','4':'ABR','5':'MAY','6':'JUN','7':'JUL','8':'AGO','9':'SEP','10':'OCT','11':'NOV','12':'DIC'}
now = datetime.datetime.now()

solicitud = pd.DataFrame(columns= ['Compania','Descripción','U.M Inventario','U.M Compras','Tipo Inv Siesa','IVA','Grupo WA','Desc. Grupo WA','Subgrupo WA','Desc. Subgrupo WA','U.N','Referencia','COD. COUPA','BIEN CAPITAL','Correos'])
driver = webdriver.Chrome("C:\Python\Python37\chromedriver")
lista_correos = []

# Carga de todas las tablas y convertirlas y limpiarlas
consecutivos_CONTEGRAL = pd.read_excel('consecutivos_CONTEGRAL.xlsx',header = 0,index_col=0)
consecutivos_CONTEGRAL = consecutivos_CONTEGRAL.replace(np.nan,"")
consecutivos_FINCA = pd.read_excel('consecutivos_FINCA.xlsx',header = 0,index_col=0)
consecutivos_FINCA = consecutivos_FINCA.replace(np.nan,"")
consecutivos_ATA = pd.read_excel('consecutivos_ATA.xlsx',header = 0,index_col=0)
consecutivos_ATA.index = consecutivos_ATA.index.map(str)
consecutivos_ATA = consecutivos_ATA.replace(np.nan,"")
consecutivos_PIC = pd.read_excel('consecutivos_PIC.xlsx',header = 0,index_col=0)
consecutivos_PIC.index = consecutivos_PIC.index.map(str)
consecutivos_PIC = consecutivos_PIC.replace(np.nan,"")
consecutivos_OPAV = pd.read_excel('consecutivos_OPAV.xlsx',header = 0,index_col=0)
consecutivos_OPAV = consecutivos_OPAV.replace(np.nan,"")
codigos_r = pd.read_excel('codigos_R.xlsx',header = 0,index_col=0)
codigos_r.index = codigos_r.index.map(str)
codigos_r = codigos_r.replace(np.nan,"")

#Estas son para ver el formato del código
estructura_codigos_CONTEGRAL = pd.read_excel('estructura_codigos_CONTEGRAL.xlsx',header = 0,index_col=0)
estructura_codigos_CONTEGRAL.index = estructura_codigos_CONTEGRAL.index.map(str)
estructura_codigos_FINCA = pd.read_excel('estructura_codigos_FINCA.xlsx',header = 0,index_col=0)
estructura_codigos_FINCA.index = estructura_codigos_FINCA.index.map(str)
estructura_codigos_ATA = pd.read_excel('estructura_codigos_ATA.xlsx',header = 0,index_col=0)
estructura_codigos_ATA.index = estructura_codigos_ATA.index.map(str)
estructura_codigos_PIC = pd.read_excel('estructura_codigos_PIC.xlsx',header = 0,index_col=0)
estructura_codigos_PIC.index = estructura_codigos_PIC.index.map(str)
estructura_codigos_OPAV = pd.read_excel('estructura_codigos_OPAV.xlsx',header = 0,index_col=0)
estructura_codigos_OPAV.index = estructura_codigos_OPAV.index.map(str)
estructura_codigos_OPAV_AF = pd.read_excel('estructura_codigos_OPAV_AF.xlsx',header = 0,index_col=0)
estructura_codigos_OPAV_AF.index = estructura_codigos_OPAV_AF.index.map(str)


def func():
    exportar_CONTEGRAL = False
    exportar_FINCA = False
    exportar_ATA = False
    exportar_PIC = False
    exportar_OPAV = False
    exportar_R = False

    driver.get('https://accounts.google.com/signin/v2/identifier?continue=https%3A%2F%2Fmail.google.com%2Fmail%2F&service=mail&sacu=1&rip=1&flowName=GlifWebSignIn&flowEntry=ServiceLogin')
    driver.find_element_by_id('identifierId').send_keys(cred_gmail.user)
    driver.find_element_by_id('identifierNext').click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="password"]/div[1]/div/div[1]/input')))
    try:
        driver.find_element_by_xpath('//*[@id="password"]/div[1]/div/div[1]/input').send_keys(cred_gmail.password)
    except:
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="password"]/div[1]/div/div[1]/input').send_keys(cred_gmail.password)
    driver.find_element_by_id('passwordNext').click()

    try:
        driver.find_element_by_link_text('Listo').click()
    except:
        pass
    
    conteo = input('Cual es el numero de la siguiente solicitud? ')
    conteo = int(conteo)

    continuar = False

    ix_solicitud = 0
    while not continuar:
        B = input('Entre al correo y descargue el archivo. Inmediatamente pulse ENTER ')
        correos = []
        # correos = ""
        # Extrae los correo que estan en la solicitud y al final se ponen en la tabla
        paso = driver.find_element_by_xpath("//*[@class='adn ads']") 
        tbody = paso.find_element_by_tag_name('tbody')
        spans = tbody.find_elements_by_tag_name('span')
        for span in spans:
            if (not str(span.get_attribute("email")) == "None") and (not str(span.get_attribute("email")) in ["parametadata@parameta.co", "parametrizaciones@grupobios.co","agaitan@parameta.co"]) :
                #correos = correos + str(span.get_attribute("email")) + "/"
                correos.append(str(span.get_attribute("email")))
                lista_correos.append(str(span.get_attribute("email")))

        archivo = max([os.path.join(r'C:\Users\Oscar\Downloads', x) for x in os.listdir(r'C:\Users\Oscar\Downloads')] , key = os.path.getctime)
        #arc = input('Nombre archivo ')
        #archivo = 'C:\\Users\\Oscar\\Downloads\\' + arc
        tabla = pd.read_excel(archivo, header = None)
        print (archivo)
        try:
            wb = openpyxl.load_workbook(archivo)
        except:
            print('No se pudo cargar el archivo')
            export = 'G:\Mi unidad\Data\Grupo Bios\Automatizacion_Items\Consecutivos SIESA\Solicitudes dia\solicitud_' + str(now).replace(" ","_").replace(':','-')[0:19] + ".xlsx"
            try:
                solicitud['IVA'] = solicitud['IVA'].astype(float).map(lambda n: '{:.0%}'.format(n))
            except:
                pass    
            solicitud.to_excel(export, header = True, encoding='utf-8', index = None)
            print('Solicitud exportada')
            
            ex = True
            while ex:
                exportar_consecutivos = input('Quiere actualizar las tablas de los consecutivos? S/N ')
                if exportar_consecutivos =='S':
                    esto_es_una_prueba = True
                    ex = False
                elif exportar_consecutivos =='N':
                    esto_es_una_prueba = False
                    ex = False

            if esto_es_una_prueba:
                if exportar_CONTEGRAL:
                    consecutivos_CONTEGRAL.to_excel('consecutivos_CONTEGRAL.xlsx',header = True, encoding = 'utf-8')
                    print('Exportado CONTEGRAL')
                if exportar_OPAV:
                    consecutivos_OPAV.to_excel('consecutivos_OPAV.xlsx',header = True, encoding = 'utf-8')
                    print('Exportado OPAV')
                if exportar_ATA:
                    consecutivos_ATA.to_excel('consecutivos_ATA.xlsx',header = True, encoding = 'utf-8')
                    print('Exportado ATA')
                if exportar_FINCA:
                    consecutivos_FINCA.to_excel('consecutivos_FINCA.xlsx',header = True, encoding = 'utf-8')
                    print('Exportado FINCA')
                if exportar_PIC:
                    consecutivos_PIC.to_excel('consecutivos_PIC.xlsx',header = True, encoding = 'utf-8')
                    print('Exportado PIC')
                if exportar_R:
                    codigos_r.to_excel('codigos_R.xlsx',header = True, encoding = 'utf-8')
                    print('Exportado codigos R')
            driver.close()

        sheet_1 = wb.get_active_sheet()
        a, b = tabla.shape
        inner = False
        ya_tiene_SIESA = False

        #Este ciclo busca el nombre de la compañia y la fila en la que estan los encabezados
        for i in range(0,a):
            for j in range(0,b):

                if tabla.iloc[i,j] in ["X",'x']:
                    compania = tabla.iloc[i,j+1]
                    print('Solicitud de la compania {}'.format(compania))

                if (str(tabla.iloc[i,j])) in ['Descripción ','Descripción','DESCRIPCION','DESCRIPCIÓN']:
                    row_desc = i
                    col_desc = j
                    inner = True
                    break
            if inner:
                break
        
        #Este ciclo busca el indicador de la columna de cada campo del item
        for j in range(col_desc,b):
            if str(tabla.iloc[row_desc,j]) == 'U.M Inventario':                
                col_um_invent = j
            
            if str(tabla.iloc[row_desc,j]) == 'U.M Compras':
                col_um_compras = j
            
            if str(tabla.iloc[row_desc,j]) == 'Tipo Inv Siesa':
                col_tipo_inv = j

            if str(tabla.iloc[row_desc,j]) == 'IVA':
                col_IVA = j

            if str(tabla.iloc[row_desc,j]) in ['Grupo WA','Grupo']:
                col_grupo_WA = j
            
            if str(tabla.iloc[row_desc,j]) in ['Desc. Grupo WA','Desc. Grupo']:
                col_desc_grupo_WA = j
            
            if str(tabla.iloc[row_desc,j]) in ['Subgrupo WA', 'Subgrupo']:
                col_subgrupo_WA = j

            if str(tabla.iloc[row_desc,j]) in ['Desc. Subgrupo WA','Desc. Subgrupo']:
                col_desc_subgrupo_WA = j

            if str(tabla.iloc[row_desc,j]) in ['U.N', 'U.N ', 'Grupo Imp Siesa']:
                col_un = j
            
            if str(tabla.iloc[row_desc,j]) in ['Referencia','Referencia UNOE']:
                col_referencia = j
            
            try:
                if str(tabla.iloc[i,j]) in ['REFERENCIA EN COUPA','Codigos Coupa','CODIGO COUPA','COD. COUPA','Codigo Cupa','Codigo Coupa','Codogo Coupa','CÓDIGO COUPA']:
                    col_COUPA = j
            except:
                continue
            
            try:
                if str(tabla.iloc[i,j]) in ['Bien de Capital','BIEN CAPITAL','Bien de Capital ','Bien Capital']:
                    col_bien_capital = j
            except:
                continue
        
        #Este ciclo va item por item buscando la informacion de cada uno. Evalua si ya tiene codigo COUPA o si ya tiene código SIESA
        ix = row_desc + 1
        for k in range(0,100):
            try:
                while not pd.isna(tabla.iloc[ix,col_desc]):
                    solicitud.loc[ix_solicitud,'Compania'] = dict_comp[compania]
                    solicitud.loc[ix_solicitud,'Descripción'] = tabla.iloc[ix,col_desc]
                    print(tabla.iloc[ix,col_desc])
                    try:
                        if tabla.iloc[ix,col_um_invent] in ['UND','UN','UNID ','UNIDAD']:
                            solicitud.loc[ix_solicitud,'U.M Inventario'] = 'UNID'
                        else:
                            solicitud.loc[ix_solicitud,'U.M Inventario'] = tabla.iloc[ix,col_um_invent]
                    except:
                        pass
                    if tabla.iloc[ix,col_um_compras] in ['UND','UN','UNID ','UNIDAD']:
                        solicitud.loc[ix_solicitud,'U.M Compras'] = 'UNID'
                    else:
                        solicitud.loc[ix_solicitud,'U.M Compras'] = tabla.iloc[ix,col_um_compras]
                    solicitud.loc[ix_solicitud,'Tipo Inv Siesa'] = tabla.iloc[ix,col_tipo_inv]
                    if solicitud.loc[ix_solicitud,'Tipo Inv Siesa'][0:2] == 'AF':
                        lista_correos.append('liliana.rojas@grupobios.co')
                        correos.append('liliana.rojas@grupobios.co')
                    solicitud.loc[ix_solicitud,'IVA'] = tabla.iloc[ix,col_IVA]
                    solicitud.loc[ix_solicitud,'Grupo WA'] = tabla.iloc[ix,col_grupo_WA]
                    solicitud.loc[ix_solicitud,'Desc. Grupo WA'] = tabla.iloc[ix,col_desc_grupo_WA]
                    solicitud.loc[ix_solicitud,'Subgrupo WA'] = tabla.iloc[ix,col_subgrupo_WA]
                    solicitud.loc[ix_solicitud,'Desc. Subgrupo WA'] = tabla.iloc[ix,col_desc_subgrupo_WA]
                    solicitud.loc[ix_solicitud,'U.N'] = tabla.iloc[ix,col_un]
                    solicitud.loc[ix_solicitud,'Correos'] = correos
                    if not pd.isna(tabla.iloc[ix,col_referencia]):
                        if str(tabla.iloc[ix,col_referencia])[0:2] in ['2B','2D','2R','2S','2U','2V','2W','3A','3F','3G','3H','3I','3K','3L','3M','3N','3O','3P','3Q','3T','5D','6Z','7E','8D','8L','8Q','8H','2X','8M','D5','8O','8F']:
                            solicitud.loc[ix_solicitud,'COD. COUPA'] = tabla.iloc[ix,col_referencia]
                            sheet_1.cell(row=ix + 1,column=col_referencia +2).value = tabla.iloc[ix,col_referencia]
                        else:
                            referencia_2 = tabla.iloc[ix,col_referencia]
                            ya_tiene_SIESA = True

                    try:
                        if not pd.isna(tabla.iloc[ix,col_COUPA]):
                            solicitud.loc[ix_solicitud,'COD. COUPA'] = tabla.iloc[ix,col_COUPA]
                    except:
                        pass
                    
                    try:
                        if not pd.isna(tabla.iloc[ix,col_bien_capital]):
                            solicitud.loc[ix_solicitud,'BIEN CAPITAL'] = tabla.iloc[ix,col_bien_capital]
                    except:
                        pass

                    #Aca se le asignaría el código SIESA dependiendo de la compania
                    if not ya_tiene_SIESA:
                        if compania == 'CONTEGRAL':
                            exportar_CONTEGRAL = True
                            if tabla.loc[ix,col_tipo_inv] == 'MP-EMPAETI':
                                if tabla.loc[ix,col_desc_subgrupo_WA] in ['SACOS POLIPROPILENO','ROLLOS X KG','BOPP (POLIP. LAMINADO)','LAMINADOS']:
                                    refer_anter = consecutivos_CONTEGRAL.loc['4XXX','consecutivo']
                                    refer_nueva = refer_anter + 1
                                    consecutivos_CONTEGRAL.loc['4XXX','consecutivo'] = refer_nueva
                                    referencia = refer_nueva
                                elif tabla.loc[ix,col_desc_subgrupo_WA] in ['ETIQUETAS ADHESIVAS','EQUIPO AVICOLA','ETIQUETAS NO ADHESIVAS','CINTAS','PVC Y PELICULAS STRETCH','EMPAQUES Y SELLOS','BOLSAS IMPRESAS','BOLSAS SIN IMPRESION','HILOS']: 
                                    refer_anter = consecutivos_CONTEGRAL.loc['9XXX','consecutivo']
                                    refer_nueva = refer_anter + 1
                                    consecutivos_CONTEGRAL.loc['9XXX','consecutivo'] = refer_nueva
                                    referencia = refer_nueva
                            else:
                                try:
                                    refer_anter = consecutivos_CONTEGRAL.loc[estructura_codigos_CONTEGRAL.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']
                                    refer_nueva = refer_anter + 1
                                    consecutivos_CONTEGRAL.loc[estructura_codigos_CONTEGRAL.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo'] = refer_nueva
                                    referencia = str(consecutivos_CONTEGRAL.loc[estructura_codigos_CONTEGRAL.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'prefijo']) + str(consecutivos_CONTEGRAL.loc[estructura_codigos_CONTEGRAL.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']) + str(consecutivos_CONTEGRAL.loc[estructura_codigos_CONTEGRAL.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'sufijo'])
                                except:
                                    print('*****************ATENCION*****************')
                                    print('El tipo de inventario ' + str(tabla.iloc[ix,col_tipo_inv]) + ' no existe')
                                    ix = ix + 1
                                    break
                        elif compania == 'FINCA':
                            if tabla.loc[ix,col_tipo_inv] == 'RP-REPU':
                                try:
                                    exportar_R = True
                                    if tabla.loc[ix,col_desc_subgrupo_WA] == 'DEPOSITO':
                                        indice = 'FERRETERIA'
                                    elif tabla.loc[ix,col_desc_subgrupo_WA] == 'COMBUSTIBLES Y LUBRICANTES':
                                        indice = 'ADITIVOS'
                                    elif tabla.loc[ix,col_desc_subgrupo_WA] in ['REPUESTOS COSEDORAS','GRANJAS E INCUBACION','REPUESTOS VEHICULOS','SISTEMAS Y EQUIPOS HIDRAULICOS','TRATAMIENTO DE AGUAS','GLP','REPUESTOS MECANICOS','REFRIGERACION']:
                                        indice = 'OTROS REPUESTOS'
                                    elif tabla.loc[ix,col_desc_subgrupo_WA] in ['EQUIPOS Y REPUESTOS DE MAQUINARIA','BANDAS TRANSPORTADORAS']:
                                        indice = 'REPUESTOS VARIOS'
                                    elif tabla.loc[ix,col_desc_subgrupo_WA] in ['EMPAQUES Y SELLOS']:
                                        indice = 'REPUESTOS EMPACADORAS Y ENFARDADORAS'
                                    else:
                                        indice = tabla.iloc[ix,col_desc_subgrupo_WA]
                                    refer_anter = codigos_r.loc[indice,'codigo']
                                    refer_nueva = refer_anter + 1
                                    codigos_r.loc[indice,'codigo'] = refer_nueva
                                    referencia = str(codigos_r.loc[indice,'prefijo']) + str(codigos_r.loc[indice,'codigo'])
                                except:
                                    print('*****************ATENCION*****************')
                                    print('El sub grupo ' + str(tabla.iloc[ix,col_desc_subgrupo_WA]) + ' no existe')
                                    ix = ix +1 
                                    break
                            elif tabla.loc[ix,col_tipo_inv] == 'MP-EMPAETI':
                                try:
                                    exportar_FINCA = True
                                    if tabla.loc[ix,col_desc_subgrupo_WA] in ['ETIQUETAS ADHESIVAS','SACOS POLIPROPILENO','CINTAS','BOLSAS SIN IMPRESION','PVC Y PELICULAS STRETCH','LAMINADOS','BOPP (POLIP.LAMINADO)','ROLLOS X KG','CORRUGADOS Y PLEGADIZAS']:
                                        refer_anter = consecutivos_FINCA.loc['40XXXX','consecutivo']
                                        refer_nueva = refer_anter + 1
                                        consecutivos_FINCA.loc['40XXXX','consecutivo'] = refer_nueva
                                        referencia =str(consecutivos_FINCA.loc['40XXXX','consecutivo'])
                                    elif tabla.loc[ix,col_desc_subgrupo_WA] in ['ETIQUETAS NO ADHESIVAS','BOLSAS IMPRESAS','HILOS','LAMINADOS','EQUIPO AVICOLA']:
                                        refer_anter = consecutivos_FINCA.loc['009XXX','consecutivo']
                                        refer_nueva = refer_anter + 1
                                        consecutivos_FINCA.loc['009XXX','consecutivo'] = refer_nueva
                                        referencia = str(consecutivos_FINCA.loc['009XXX','consecutivo'])
                                except:
                                    print('*****************ATENCION*****************')
                                    print('El tsub grupo ' + str(tabla.iloc[ix,col_desc_subgrupo_WA]) + ' no existe')
                                    ix = ix +1 
                                    break
                            else:
                                try:
                                    exportar_FINCA = True
                                    refer_anter = consecutivos_FINCA.loc[estructura_codigos_FINCA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']
                                    refer_nueva = refer_anter + 1
                                    consecutivos_FINCA.loc[estructura_codigos_FINCA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo'] = refer_nueva
                                    referencia = str(consecutivos_FINCA.loc[estructura_codigos_FINCA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'prefijo']) + str(consecutivos_FINCA.loc[estructura_codigos_FINCA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']) + str(consecutivos_FINCA.loc[estructura_codigos_FINCA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'sufijo'])
                                except:
                                    print('*****************ATENCION*****************')
                                    print('El tipo de inventario ' + str(tabla.iloc[ix,col_tipo_inv]) + ' no existe')
                                    ix = ix +1 
                                    break
                        elif compania == 'AVICOLA TRIPLE A':
                            try:
                                exportar_ATA = True
                                refer_anter = consecutivos_ATA.loc[estructura_codigos_ATA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']
                                refer_nueva = refer_anter + 1
                                consecutivos_ATA.loc[estructura_codigos_ATA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo'] = refer_nueva
                                referencia = str(consecutivos_ATA.loc[estructura_codigos_ATA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'prefijo']) + str(consecutivos_ATA.loc[estructura_codigos_ATA.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']) 
                            except:
                                print('*****************ATENCION*****************')
                                print('El tipo de inventario ' + str(tabla.iloc[ix,col_tipo_inv])+ ' no existe')
                                ix = ix + 1
                                break
                        elif compania in ['PIC','PIC ']:
                            if tabla.iloc[ix,col_desc_subgrupo_WA] == 'ALIMENTO BALANCEADO':
                                print('Es de alimento balanceado para PIC')
                                break
                            try:
                                exportar_PIC = True
                                refer_anter = consecutivos_PIC.loc[estructura_codigos_PIC.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']
                                refer_nueva = refer_anter + 1
                                consecutivos_PIC.loc[estructura_codigos_PIC.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo'] = refer_nueva
                                referencia = str(consecutivos_PIC.loc[estructura_codigos_PIC.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'prefijo'])  + str(consecutivos_PIC.loc[estructura_codigos_PIC.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']) 
                            except:
                                print('*****************ATENCION*****************')
                                print('El tipo de inventario ' + str(tabla.iloc[ix,col_tipo_inv])+ ' no existe')
                                ix = ix + 1
                                break
                        elif compania == 'OPAV':
                            exportar_OPAV = True
                            if tabla.iloc[ix,col_tipo_inv][0:2] =='AF':
                                try:
                                    refer_anter = consecutivos_OPAV.loc[estructura_codigos_OPAV_AF.loc[tabla.iloc[ix,col_tipo_inv],tabla.iloc[ix,col_un]],'consecutivo']
                                    refer_nueva = refer_anter + 1
                                    consecutivos_OPAV.loc[estructura_codigos_OPAV_AF.loc[tabla.iloc[ix,col_tipo_inv],tabla.iloc[ix,col_un]],'consecutivo'] = refer_nueva
                                    referencia = refer_nueva
                                except:
                                    print('*****************ATENCION*****************')
                                    print('El tipo de inventario ' + str(tabla.iloc[ix,col_tipo_inv])+ ' no existe')
                                    ix = ix + 1
                                    break
                            else:
                                try:
                                    refer_anter = consecutivos_OPAV.loc[estructura_codigos_OPAV.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo']
                                    refer_nueva = refer_anter + 1
                                    consecutivos_OPAV.loc[estructura_codigos_OPAV.loc[tabla.iloc[ix,col_tipo_inv],'estructura_codigo'],'consecutivo'] = refer_nueva
                                    referencia = refer_nueva
                                except:
                                    print('*****************ATENCION*****************')
                                    print('El tipo de inventario ' + str(tabla.iloc[ix,col_tipo_inv])+ ' no existe')
                                    ix = ix + 1
                                    break
                    if ya_tiene_SIESA:
                        solicitud.loc[ix_solicitud,'Referencia'] = referencia_2 
                    else:
                        solicitud.loc[ix_solicitud,'Referencia'] = referencia
                        sheet_1.cell(row=ix +1 ,column=col_referencia + 1).value = referencia
                    ix_solicitud = ix_solicitud + 1
                    ix = ix+1
                    ya_tiene_SIESA = False
            except IndexError:
                break
        if "felipe.betancur@grupobios.co " in correos:
            solicitante = "\FB"
        elif "jaime.lopez@grupobios.co" in correos:
            solicitante = "\JAL"
        elif "paula.hoyos@grupobios.co" in correos:
            solicitante = "\PAH"
        elif "mary.calderon@grupobios.co" in correos:
            solicitante = "\MJC"
        elif "catalina.mejia@grupobios.co" in correos:
            solicitante = "\CM"
        else:
            solicitante = "\\"

        sol_dia = 'G:\Mi unidad\DATA\Grupo Bios\Administración de ítems\Solicitudes de creación Siesa'+ solicitante + " " + dict_mes[str(now.month)] + " " + str(now.day) + " " + compania[0:1] + " " + str(conteo) + ".xlsx"
        wb.save(sol_dia)
        solicitante = ""
        del(col_um_invent,col_um_compras,col_tipo_inv,col_IVA,col_grupo_WA,col_desc_grupo_WA,col_subgrupo_WA,col_desc_subgrupo_WA,col_un)
        
        # Si se encontró la columna...eliminela. Hay formatos que no tienen
        try:
            del(col_COUPA)
        except:
            pass

        try:
            del(col_bien_capital)
        except:
            pass

        driver.find_element_by_xpath('//*[@class="aic"]/div/div').click()
        try:
            driver.find_element_by_xpath('//*[@class="eV"]/div/div/textarea').send_keys('parametrizaciones@grupobios.co')
        except:
            time.sleep(2)
            driver.find_element_by_xpath('//*[@class="eV"]/div/div/textarea').send_keys('parametrizaciones@grupobios.co')
        driver.find_element_by_xpath('//*[@class="aA6"]/span/span/span').click()

        for correo in correos:
            driver.find_element_by_xpath('//*[@class="GS"]/tbody/tr[2]/td[2]/div/div/textarea').send_keys(correo + " ")

        driver.find_element_by_name('subjectbox').send_keys('Solicitud creación de items {} {}'.format(compania,str(conteo)))
        conteo = conteo + 1
        driver.find_element_by_xpath('//*[@class="Am Al editable LW-avf"]').send_keys('Buen dia')
        driver.find_element_by_xpath('//*[@class="Am Al editable LW-avf"]').send_keys(Keys.ENTER)
        driver.find_element_by_xpath('//*[@class="Am Al editable LW-avf"]').send_keys(Keys.ENTER)
        driver.find_element_by_xpath('//*[@class="Am Al editable LW-avf"]').send_keys('Solicito su colaboración con la creación de los items adjuntos a la compañía {}'.format(compania))
        #driver.find_element_by_class_name('Hp').click()
        #driver.find_element_by_class_name('Ha').click()

        #Este bucle sirve para determinar si se van a cargar mas items o para acabar la compilacion
        c = False
        while not c:
            cont = input('Hay mas items? S/N ')
            if cont in ['S','N']:
                c = True
                if cont == 'S':
                    continue
                elif cont == 'N':
                    continuar = True
    
    # Estas dos lineas exportan la solicitud
    export = 'G:\Mi unidad\Data\Grupo Bios\Automatizacion_Items\Consecutivos SIESA\Solicitudes dia\solicitud_' + str(now).replace(" ","_").replace(':','-')[0:19] + ".xlsx"
    try:
        solicitud['IVA'] = solicitud['IVA'].astype(float).map(lambda n: '{:.0%}'.format(n))
    except:
        pass    
    solicitud.to_excel(export, header = True, encoding='utf-8', index = None)
    print('Solicitud exportada')    

    final_list = [] 
    for num in lista_correos: 
        if num not in final_list: 
            final_list.append(num)
    
    #Aca se exporta cada una de las tablas, dependiendo si se utilizaron
    ex = True
    while ex:
        exportar_consecutivos = input('Quiere actualizar las tablas de los consecutivos? S/N ')
        if exportar_consecutivos =='S':
            esto_es_una_prueba = True
            ex = False
        elif exportar_consecutivos =='N':
            esto_es_una_prueba = False
            ex = False

    if esto_es_una_prueba:
        if exportar_CONTEGRAL:
            consecutivos_CONTEGRAL.to_excel('consecutivos_CONTEGRAL.xlsx',header = True, encoding = 'utf-8')
            print('Exportado CONTEGRAL')
        if exportar_OPAV:
            consecutivos_OPAV.to_excel('consecutivos_OPAV.xlsx',header = True, encoding = 'utf-8')
            print('Exportado OPAV')
        if exportar_ATA:
            consecutivos_ATA.to_excel('consecutivos_ATA.xlsx',header = True, encoding = 'utf-8')
            print('Exportado ATA')
        if exportar_FINCA:
            consecutivos_FINCA.to_excel('consecutivos_FINCA.xlsx',header = True, encoding = 'utf-8')
            print('Exportado FINCA')
        if exportar_PIC:
            consecutivos_PIC.to_excel('consecutivos_PIC.xlsx',header = True, encoding = 'utf-8')
            print('Exportado PIC')
        if exportar_R:
            codigos_r.to_excel('codigos_R.xlsx',header = True, encoding = 'utf-8')
            print('Exportado codigos R')
    driver.close()
        

if __name__ == "__main__":
    func()